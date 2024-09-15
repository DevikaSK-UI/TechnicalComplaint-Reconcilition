import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, Scrollbar
import os
import webbrowser

def load_file(prompt):
    """Load an Excel file using a file dialog."""
    return filedialog.askopenfilename(title=prompt, filetypes=[("Excel files", "*.xlsx")])

def perform_comparison(ccc_path, edc_path, output_path):
    # Load the datasets
    ccc_df = pd.read_excel(ccc_path)
    edc_df = pd.read_excel(edc_path)

    # Define relevant columns including the new Trial/Study Number column
    ccc_columns = ['Subject/Patient ID', 'Technical Complaint No.', 'AE related', 'DUN Number', 'Trial/Study Number']
    edc_columns = ['Subject', 'Seq No', 'AE related', 'Dispense Unit Number ID', 'Trial/Study Number']

    # Ensure all relevant columns are present
    ccc_df = ccc_df[ccc_columns]
    edc_df = edc_df[edc_columns]

    # Rename CCC columns to match EDC columns for comparison
    ccc_df = ccc_df.rename(columns={
        'Subject/Patient ID': 'Subject',
        'Technical Complaint No.': 'Seq No',
        'DUN Number': 'Dispense Unit Number ID'
    })

    # Initialize the status column and mismatch details
    edc_df['Status'] = 'Not Present'
    edc_df['Mismatch_Details'] = ''

    # Perform the comparison for EDC rows
    for i, row in edc_df.iterrows():
        ccc_row = ccc_df.loc[ccc_df['Subject'] == row['Subject']]
        if not ccc_row.empty:
            mismatch_details = []
            # Compare each relevant column and add detailed mismatch data
            if row['Seq No'] != ccc_row.iloc[0]['Seq No']:
                mismatch_details.append(f"Seq No (EDC: {row['Seq No']} / CCC: {ccc_row.iloc[0]['Seq No']})")
            if row['AE related'] != ccc_row.iloc[0]['AE related']:
                mismatch_details.append(f"AE related (EDC: {row['AE related']} / CCC: {ccc_row.iloc[0]['AE related']})")
            if row['Dispense Unit Number ID'] != ccc_row.iloc[0]['Dispense Unit Number ID']:
                mismatch_details.append(f"Dispense Unit Number ID (EDC: {row['Dispense Unit Number ID']} / CCC: {ccc_row.iloc[0]['Dispense Unit Number ID']})")
            if row['Trial/Study Number'] != ccc_row.iloc[0]['Trial/Study Number']:
                mismatch_details.append(f"Trial/Study Number (EDC: {row['Trial/Study Number']} / CCC: {ccc_row.iloc[0]['Trial/Study Number']})")
            
            if not mismatch_details:
                edc_df.at[i, 'Status'] = 'Match'
            else:
                edc_df.at[i, 'Status'] = 'Mismatch'
                edc_df.at[i, 'Mismatch_Details'] = ', '.join(mismatch_details)
        else:
            edc_df.at[i, 'Status'] = 'Not Present'
            edc_df.at[i, 'Mismatch_Details'] = 'Not Present in CCC'

    # Identify rows in CCC_TC not present in EDC_TC
    not_in_edc = []
    for i, row in ccc_df.iterrows():
        edc_row = edc_df.loc[edc_df['Subject'] == row['Subject']]
        if edc_row.empty:
            not_in_edc.append(row)

    # Convert the missing rows to DataFrame
    not_in_edc_df = pd.DataFrame(not_in_edc, columns=['Subject', 'Seq No', 'AE related', 'Dispense Unit Number ID', 'Trial/Study Number'])
    not_in_edc_df['Status'] = 'Not Present'
    not_in_edc_df['Mismatch_Details'] = 'Not Present in EDC'

    # Append missing rows to EDC DataFrame with status 'Not Present'
    if not not_in_edc_df.empty:
        edc_df = pd.concat([edc_df, not_in_edc_df], ignore_index=True)

    # Update column names with the original dataset column names separated by a slash
    final_columns = {
        'Subject': 'Subject/Patient ID / Subject',
        'Seq No': 'Technical Complaint No. / Seq No',
        'Dispense Unit Number ID': 'DUN Number / Dispense Unit Number ID',
        'AE related': 'AE related / AE related',
        'Trial/Study Number': 'Trial/Study Number / Trial/Study Number',
        'Status': 'Status',
        'Mismatch_Details': 'Mismatch_Details'
    }

    # Rename columns in edc_df to reflect the original dataset columns
    edc_df = edc_df.rename(columns=final_columns)

    # Select only the relevant columns for the final output
    edc_df = edc_df[list(final_columns.values())]

    # Save the updated DataFrame to Excel with the provided filename
    edc_df.to_excel(output_path, index=False)

    # Load the workbook and select the active sheet
    wb = load_workbook(output_path)
    ws = wb.active

    # Define colors for fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    Amber_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Define border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply borders, font style, and alignment to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply font styles and color fill for headers
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill

    # Adjust the width of the columns based on the content
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Apply color formatting based on status
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value == 'Mismatch':
                cell.fill = red_fill
            elif cell.value == 'Match':
                cell.fill = green_fill
            elif cell.value == 'Not Present':
                cell.fill = Amber_fill

    # Save the workbook with the updated filename
    wb.save(output_path)
    return output_path

def open_output_file(output_path):
    """Open the output file in the default application."""
    webbrowser.open(output_path)

def select_ccc_file():
    """Handle CCC file selection."""
    file_path = load_file("Select the CCC file")
    if file_path:
        ccc_path.set(file_path)
        ccc_label.config(text=f"CCC File: {os.path.basename(file_path)}")

def select_edc_file():
    """Handle EDC file selection."""
    file_path = load_file("Select the EDC file")
    if file_path:
        edc_path.set(file_path)
        edc_label.config(text=f"EDC File: {os.path.basename(file_path)}")

def start_comparison():
    """Start the comparison process."""
    if ccc_path.get() and edc_path.get():
        output_filename = filedialog.asksaveasfilename(
            title="Save Output File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{datetime.now().strftime('%Y-%m-%d')}_Technical_Reconciliation.xlsx"
        )
        if output_filename:
            try:
                global output_path
                output_path = perform_comparison(ccc_path.get(), edc_path.get(), output_filename)
                messagebox.showinfo("Comparison Complete", f"Comparison completed successfully!\nOutput saved at: {output_path}")

                # Enable "Show Results Here" button and "Open Output File" button
                show_results_button.config(state=tk.NORMAL)
                open_file_button.config(state=tk.NORMAL)
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {str(e)}")
    else:
        messagebox.showwarning("Input Error", "Please select both CCC and EDC files.")

def show_results():
    """Display results in a Tkinter Text widget with color coding."""
    if not output_path:
        messagebox.showwarning("No Results", "Please perform a comparison first.")
        return

    # Read the result from the Excel file
    df = pd.read_excel(output_path)

    # Clear previous content in text widget
    result_text.delete(1.0, tk.END)

    # Add header
    headers = df.columns.tolist()
    column_widths = {header: len(header) for header in headers}

    # Find the maximum width for each column
    for _, row in df.iterrows():
        for header in headers:
            column_widths[header] = max(column_widths[header], len(str(row[header])))

    # Add header to the text widget
    header_line = " | ".join(header.ljust(column_widths[header]) for header in headers) + "\n"
    result_text.insert(tk.END, header_line)

    # Configure tags for color coding
    result_text.tag_configure("red", foreground="white", background="red")
    result_text.tag_configure("green", foreground="white", background="green")
    result_text.tag_configure("amber", foreground="white", background="#FFBF00") 

    # Add data rows
    for index, row in df.iterrows():
        row_text = " | ".join(str(row[header]).ljust(column_widths[header]) for header in headers) + "\n"
        result_text.insert(tk.END, row_text)

        # Apply color coding to the row based on Status
        if row['Status'] == 'Mismatch':
            result_text.tag_add("red", f"{index+2}.0", f"{index+2}.end")
        elif row['Status'] == 'Match':
            result_text.tag_add("green", f"{index+2}.0", f"{index+2}.end")
        elif row['Status'] == 'Not Present':
            result_text.tag_add("amber", f"{index+2}.0", f"{index+2}.end")

# Create the main window
root = tk.Tk()
root.title("Technical Complaint Reconciliation")

# Set window size and background color
root.geometry("1200x600")  # Adjusted size to accommodate result display
root.configure(bg="darkblue")

# Add a label in a big white box for the welcome text
welcome_frame = tk.Frame(root, bg="white", padx=10, pady=10)
welcome_frame.pack(pady=10, fill=tk.X)

welcome_label = tk.Label(welcome_frame, text="Welcome to the Technical Complaint Reconciliation Tool", font=("Arial", 26, "bold"), bg="white", fg="darkblue")
welcome_label.pack()

# Create variables to hold file paths
ccc_path = tk.StringVar()
edc_path = tk.StringVar()

# Buttons for CCC selection
select_ccc_button = tk.Button(root, text="Select CCC File", command=select_ccc_file, height=2, width=25, bg="white", fg="darkblue")
select_ccc_button.pack(pady=10)

# Labels to display selected file paths
ccc_label = tk.Label(root, text="CCC File: Not selected", bg="darkblue", fg="white", font=("Arial", 12))
ccc_label.pack(pady=5)

# Buttons for EDC selection
select_edc_button = tk.Button(root, text="Select EDC File", command=select_edc_file, height=2, width=25, bg="white", fg="darkblue")
select_edc_button.pack(pady=10)

# Labels to display selected file paths
edc_label = tk.Label(root, text="EDC File: Not selected", bg="darkblue", fg="white", font=("Arial", 12))
edc_label.pack(pady=5)

# Button to start comparison
compare_button = tk.Button(root, text="Compare the Excel Sheets", command=start_comparison, height=2, width=25, bg="#FFC0C0", fg="darkblue")
compare_button.pack(pady=10)

# Button to show results in the GUI
show_results_button = tk.Button(root, text="Show Results Here", command=show_results, height=2, width=25, bg="#87CEEB", fg="darkblue", state=tk.DISABLED)
show_results_button.pack(pady=10)

# Button to open the output file
open_file_button = tk.Button(root, text="Open Output File", command=lambda: open_output_file(output_path), height=2, width=25, bg="lightgreen", fg="darkblue", state=tk.DISABLED)
open_file_button.pack(pady=10)

# Frame to contain the Text widget and scrollbars
text_frame = tk.Frame(root)
text_frame.pack(expand=True, fill=tk.BOTH)

# Text widget for displaying results
result_text = tk.Text(text_frame, wrap=tk.NONE, bg="white", fg="black", height=15, width=150)
result_text.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

# Scrollbars
vertical_scrollbar = Scrollbar(text_frame, orient=tk.VERTICAL, command=result_text.yview)
vertical_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

horizontal_scrollbar = Scrollbar(text_frame, orient=tk.HORIZONTAL, command=result_text.xview)
horizontal_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

# Configure the Text widget to use the scrollbars
result_text.config(yscrollcommand=vertical_scrollbar.set, xscrollcommand=horizontal_scrollbar.set)

# Global variable for the output path
output_path = None

# Start the GUI event loop
root.mainloop()

